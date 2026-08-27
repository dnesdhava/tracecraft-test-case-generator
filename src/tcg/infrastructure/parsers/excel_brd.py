from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from tcg.domain.models import (
    Priority,
    Requirement,
    SourceExtraction,
    SourceLocation,
    SourceType,
)


class ExcelBRDParser:
    """Parse the structured Excel BRD adapter used by the sample corpus."""

    source_types = (SourceType.BRD_EXCEL,)
    required_sheets = {
        "Business_Requirements",
        "Functional_Requirements",
        "Business_Rules",
        "Validation_Rules",
        "Test_Data",
    }

    def accepts(self, source_type: SourceType) -> bool:
        return source_type in self.source_types

    def parse(self, path: Path, source_id: str) -> SourceExtraction:
        warnings: list[str] = []
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (OSError, ValueError, KeyError) as exc:
            raise ValueError(f"Unable to read Excel BRD: {path.name}") from exc

        missing_sheets = sorted(self.required_sheets - set(workbook.sheetnames))
        if missing_sheets:
            warnings.append(f"Missing expected worksheets: {', '.join(missing_sheets)}")
        if "Business_Requirements" not in workbook.sheetnames:
            return SourceExtraction(
                source_id,
                SourceType.BRD_EXCEL,
                warnings=tuple(warnings + ["Business_Requirements worksheet is required"]),
            )

        sheet = workbook["Business_Requirements"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return SourceExtraction(
                source_id,
                SourceType.BRD_EXCEL,
                warnings=tuple(warnings + ["Business_Requirements worksheet is empty"]),
            )
        header_map = self._headers(rows[0])
        requirements: list[Requirement] = []
        for row_number, values in enumerate(rows[1:], start=2):
            record = {
                name: self._text(values[index])
                for name, index in header_map.items()
                if index < len(values)
            }
            requirement_id = record.get("requirement id", "").strip()
            if not requirement_id:
                if any(record.values()):
                    warnings.append(f"Row {row_number} has no Requirement ID and was skipped")
                continue
            priority = self._priority(record.get("priority", ""))
            dependencies = tuple(
                item.strip()
                for item in re.split(r"[;|,]", record.get("dependencies", ""))
                if item.strip()
            )
            location = SourceLocation(
                source_id=source_id,
                label=f"Business_Requirements/{requirement_id}",
                sheet_name="Business_Requirements",
                row_number=row_number,
            )
            requirements.append(
                Requirement(
                    requirement_id=requirement_id,
                    description=record.get("requirement description", ""),
                    business_process=record.get("business process", ""),
                    functional_requirement=record.get("functional requirement", ""),
                    business_rule=record.get("business rule", ""),
                    input_field=record.get("input field", ""),
                    validation=record.get("validation", ""),
                    expected_behaviour=record.get("expected behaviour", ""),
                    priority=priority,
                    dependencies=dependencies,
                    source_id=source_id,
                    location=location,
                    confidence=1.0,
                    original_text=record.get("requirement description", ""),
                )
            )
        if not requirements:
            warnings.append("No identifiable business requirements were found")
        return SourceExtraction(
            source_id=source_id,
            source_type=SourceType.BRD_EXCEL,
            requirements=tuple(requirements),
            warnings=tuple(warnings),
        )

    @staticmethod
    def _headers(row: tuple[object, ...]) -> dict[str, int]:
        result: dict[str, int] = {}
        for index, value in enumerate(row):
            name = " ".join(str(value or "").strip().lower().split())
            if name:
                result[name] = index
        return result

    @staticmethod
    def _text(value: object) -> str:
        return " ".join(str(value or "").strip().split())

    @staticmethod
    def _priority(value: str) -> Priority:
        normalized = value.strip().upper()
        return {
            "CRITICAL": Priority.CRITICAL,
            "HIGH": Priority.HIGH,
            "MEDIUM": Priority.MEDIUM,
            "LOW": Priority.LOW,
        }.get(normalized, Priority.UNKNOWN)
